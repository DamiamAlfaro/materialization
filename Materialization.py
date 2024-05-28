import pandas as pd
import os
from openpyxl import load_workbook
from datetime import datetime
import matplotlib.pyplot as plt
import numpy as np
import json

"""
The function that starts it all.
"""
def beginning(file1,file2,file3,file4):
  print("\nMATHEMATICS | PROGRAMMING | PROFESSION | FINANCE\n")

  files = [file1,file2,file3,file4]
        
  sidesofsquare = {"MATHEMATICS":mathematics,
                   "PROGRAMMING":programming,
                   "PROFESSION":profession,
                   "FINANCE":finance}
                        
  choice = input("Side: ").upper()
  
  if choice in sidesofsquare:
      pass
  else:
    print("\nAre you cognitively impaired?...\n")
    return
 
  keys_list = list(sidesofsquare.keys()).index(choice)
  sidesofsquare[choice](choice,files[keys_list])

"""
writing() will be used to record instances of points of time
"""
def writing(self,action):
 excel_file = self.file

 book = load_workbook(excel_file)

  # Getting instance of point of time
 now = datetime.now()

 df = pd.DataFrame({
     "Year": [now.year],
     "Month": [now.month],
     "Day": [now.day],
     "Hour":[now.hour],
     "Minute": [now.minute]})

 if action not in book.sheetnames:
     ws = book.create_sheet(action)
  
 else:
     ws = book[action]

 max_row = ws.max_row

 if all(ws.cell(row=max_row,column=col).value is None for col in range(1,6)):
     pass
 else:
     max_row += 1

 for index, row in df.iterrows():
     for col_num, value in enumerate(row):
         ws.cell(row=max_row + index, column=col_num+1, value=value)

 book.save(excel_file)
 book.close()

"""
In case of unwanted files polluting the folder.
"""
def delete_ds_store_files(directory):
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file == '.DS_Store':
                file_path = os.path.join(root, file)
                try:
                    os.remove(file_path)
                    print(f'Deleted: {file_path}')
                except Exception as e:
                    print(f'Error deleting {file_path}: {e}')



"""
--------------------
Mathematics Quadrant
--------------------
"""
class mathematics():
  def __init__(self,choice,file):

    self.choice = choice
    self.file = file
    print("\nOUTSET | HALT | VIEW | MILESTONE | TESTING\n")

    mathematicsSides = {"OUTSET":self.Outset,
                        "HALT":self.Halt,
                        "VIEW":self.View,
                        "MILESTONE":self.Milestone,
                        "TESTING":self.Test}

    mathematics_choice = input(f"{choice}-Side: ").upper()

    if mathematics_choice in mathematicsSides:
      mathematicsSides[mathematics_choice]()
    else:
      print("\nOh, I see, you are a dumbass...\n")

  """
  Annotates the beginning instance time of the Quadrant
  """
  def Outset(self):
    action = "OUTSET"
    writing(self,action)

  """
  Annotates the final instance time of the Quadrant
  """
  def Halt(self):
    action = "HALT"
    writing(self,action)


  """
  Portrays the amount of hours spend on the Quadrant, and a
  Cartesian Plane (only the positive values), or also known as
  a scatter plot of the amount of time (in hours) with respect to
  the time of the day the quadrant ended.
  """
  def View(self):
    # I. First we read the excel that stores the parameters
    view_outsets = pd.read_excel(self.file,sheet_name=0,header=None)
    view_halts = pd.read_excel(self.file,sheet_name=1,header=None)
    view_accuracy = pd.read_excel(self.file,sheet_name=4,header=None)

    # II. We prepare to compare by listing all instances
    outsets = view_outsets.values.tolist()
    halts = view_halts.values.tolist()
    accuracy = view_accuracy.values.tolist()
    

    # III. Correlating the accuracy with the dates
    accuracy_dates = [date[:-1] for date in accuracy]
    accuracy_rates = [rate[3] for rate in accuracy]

    datesfloatForm = []
    for p in range(len(halts)):
       firstthreeDates = halts[p][:-2]
       comparison = []
       for q in firstthreeDates:
          comparison.append(float(q))
       datesfloatForm.append(comparison)
    

    countingAccuracies = []
    for instance in datesfloatForm:
      if instance in accuracy_dates:
        instanceIndex = accuracy_dates.index(instance)
        countingAccuracies.append(accuracy_rates[instanceIndex])
      else:
        countingAccuracies.append(0)

    # IV. Compare hours and minutes of outsets and halts for further ploting
    xaxis = []
    yaxis = []

    for x, y in zip(outsets,halts):
       hour = float(y[3]-x[3])
       minute = round((y[4]/60)-(x[4]/60),3)
       hours = hour + minute
       xaxis.append(round(hours,3))
       yaxis.append(round(y[3]+(x[3]/60),3))

    totalhours = round(sum(xaxis),2)
      
    # V. Now we need the x and y axis; i.e hours, time of the day respectively
    a = np.array(xaxis)
    b = np.array(yaxis)
    colors = np.array(countingAccuracies)
    plt.scatter(a,b,c=colors,cmap='plasma')
    plt.colorbar()
    plt.title(f"{self.choice.title()} Total Hours {totalhours}")
    plt.xlabel("Hours")
    plt.ylabel("Ending Day Time")
    plt.show()

  """
  Portrays in a timeline all the milestones accomplished
  within the Quadrant
  """
  def Milestone(self):
    # I. Transport data into proper format
    df = pd.read_excel(self.file,sheet_name=3,header=None)
    sector = self.choice
    datesOrigin = df.iloc[:,0].tolist()
    milestonesOrigin = df.iloc[:,1].tolist()
    print("\nNEW | VIEW\n")
    action = input("-> ")
    match str(action).upper():
        case "NEW":
          print("\nInput Date as: YYYY-MM-DD\n")
          date = str(input("-> "))
          try:
              datetime.strptime(date,"%Y-%m-%d")
          except:
              print("\nAre you retarded?\n")
              return
          print("\nWhat occurred that date?\n")
          milestoneInstance = input("-> ")
          milestoneAnnotation = [date,milestoneInstance]
          workbook = load_workbook(self.file)
          sheet_name = "MILESTONE"
          if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
          lastrowFirst = sheet.max_row
          lastrowSecond = sheet.max_row
          nextRow = max(lastrowFirst,lastrowSecond) + 1
          sheet.cell(row=nextRow,column=1,value=milestoneAnnotation[0])
          sheet.cell(row=nextRow,column=2,value=milestoneAnnotation[1])
          workbook.save(self.file)
            

        case "VIEW":
          df = pd.read_excel(self.file,sheet_name=3,header=None)
          dates = datesOrigin
          milestone = milestonesOrigin
          setup = pd.DataFrame(data={"Date": dates,"Milestone": milestone})
          setup["Date"] = pd.to_datetime(setup["Date"])
          setup["Level"] = [np.random.randint(-6,-2) if (i%2)==0 else np.random.randint(2,6) for i in
                                range(len(setup))]
          with plt.style.context("fivethirtyeight"):
              fig, ax = plt.subplots(figsize=(7,10))
              ax.plot([0,]* len(setup), setup.Date,"-o",color="black",markerfacecolor="white");
              ax.set_yticks(pd.date_range("2024-1-1","2024-12-30",freq="YS"), range(2024,2024));
              ax.set_xlim(-7,7);
              for idx in range(len(setup)):
                  dt,product,level = setup["Date"][idx],setup["Milestone"][idx],setup["Level"][idx]
                  dt_str = dt.strftime("%b-%Y")
                  ax.annotate(dt_str+"\n"+product,xy=(0.1 if level>0 else -0.1,dt),
                              xytext=(level,dt),
                              arrowprops=dict(arrowstyle="-",color="red",linewidth=0.8),
                              va="center",fontsize=8);
                  ax.spines[["left", "top", "right", "bottom"]].set_visible(False);
                  ax.spines[["left"]].set_position(("axes", 0.5));
                  ax.xaxis.set_visible(False);                      
                  ax.set_title(f"{sector.title()} Milestones", pad=10, loc="left",
                                    fontsize=25, fontweight="bold");
                  ax.grid(False)
          plt.show() 
    
  """
  Math Quadrant Feature: Tests our accuracy of a chapter
  and annotates it for future review.
  """
  def Test(self):
    excel = self.file
    totalQuestions = int(input("\nTotal Questions: "))
    rightQuestions = int(input("\nTotal Questions Answered Correctly: "))
    accuracyRate = round(float(rightQuestions/totalQuestions),3)
    year,month,day = datetime.now().year,datetime.now().month,datetime.now().day
    annotate = [year,month,day,accuracyRate]
    workbook = load_workbook(excel)
    sheet_name = "TESTING"
    if sheet_name in workbook.sheetnames:
      sheet = workbook[sheet_name]
    lastrowFirst = sheet.max_row
    lastrowSecond = sheet.max_row
    nextRow = max(lastrowFirst,lastrowSecond) + 1
    sheet.cell(row=nextRow,column=1,value=annotate[0])
    sheet.cell(row=nextRow,column=2,value=annotate[1])
    sheet.cell(row=nextRow,column=3,value=annotate[2])
    sheet.cell(row=nextRow,column=4,value=annotate[3])
    workbook.save(self.file)

     
    

"""
--------------------
Programming Quadrant
--------------------
"""
class programming():
  def __init__(self,choice,file):
    self.choice = choice
    self.file = file
      
    print("\nOUTSET | HALT | VIEW | MILESTONE\n")
    programmingSides = {"OUTSET":self.Outset,
                        "HALT":self.Halt,
                        "VIEW":self.View,
                        "MILESTONE":self.Milestone}
    programming_choice = input(f"{choice}-Side: ").upper()

    if programming_choice in programmingSides:
        programmingSides[programming_choice]()
    else:
        print("\nI recognize a dumbass when I see one...\n")


  def Outset(self):
    action = "OUTSET"
    writing(self,action)

  def Halt(self):
    action = "HALT"
    writing(self,action)

  def View(self):
    # I. First we read the excel that stores the parameters
    view_outsets = pd.read_excel(self.file,sheet_name=0,header=None)
    view_halts = pd.read_excel(self.file,sheet_name=1,header=None)

    # II. We prepare to compare by listing all instances
    outsets = view_outsets.values.tolist()
    halts = view_halts.values.tolist()
    

    datesfloatForm = []
    for p in range(len(halts)):
       firstthreeDates = halts[p][:-2]
       comparison = []
       for q in firstthreeDates:
          comparison.append(float(q))
       datesfloatForm.append(comparison)
    

    # IV. Compare hours and minutes of outsets and halts for further ploting
    xaxis = []
    yaxis = []

    for x, y in zip(outsets,halts):
       hour = float(y[3]-x[3])
       minute = round((y[4]/60)-(x[4]/60),3)
       hours = hour + minute
       xaxis.append(round(hours,3))
       yaxis.append(round(y[3]+(x[3]/60),3))

    totalhours = round(sum(xaxis),2)
      
    # V. Now we need the x and y axis; i.e hours, time of the day respectively
    a = np.array(xaxis)
    b = np.array(yaxis)
    plt.scatter(a,b)
    plt.title(f"{self.choice.title()} Total Hours: {totalhours}")
    plt.xlabel("Hours")
    plt.ylabel("Ending Day Time")
    plt.show()
  
  def Milestone(self):
    # I. Transport data into proper format
    df = pd.read_excel(self.file,sheet_name=3,header=None)
    sector = self.choice
    datesOrigin = df.iloc[:,0].tolist()
    milestonesOrigin = df.iloc[:,1].tolist()
    print("\nNEW | VIEW\n")
    action = input("-> ")
    match str(action).upper():
        case "NEW":
          print("\nInput Date as: YYYY-MM-DD\n")
          date = str(input("-> "))
          try:
              datetime.strptime(date,"%Y-%m-%d")
          except:
              print("\nAre you retarded?\n")
              return
          print("\nWhat occurred that date?\n")
          milestoneInstance = input("-> ")
          milestoneAnnotation = [date,milestoneInstance]
          workbook = load_workbook(self.file)
          sheet_name = "MILESTONE"
          if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
          lastrowFirst = sheet.max_row
          lastrowSecond = sheet.max_row
          nextRow = max(lastrowFirst,lastrowSecond) + 1
          sheet.cell(row=nextRow,column=1,value=milestoneAnnotation[0])
          sheet.cell(row=nextRow,column=2,value=milestoneAnnotation[1])
          workbook.save(self.file)
            

        case "VIEW":
          df = pd.read_excel(self.file,sheet_name=3,header=None)
          dates = datesOrigin
          milestone = milestonesOrigin
          setup = pd.DataFrame(data={"Date": dates,"Milestone": milestone})
          setup["Date"] = pd.to_datetime(setup["Date"])
          setup["Level"] = [np.random.randint(-6,-2) if (i%2)==0 else np.random.randint(2,6) for i in
                                range(len(setup))]
          with plt.style.context("fivethirtyeight"):
              fig, ax = plt.subplots(figsize=(7,10))
              ax.plot([0,]* len(setup), setup.Date,"-o",color="black",markerfacecolor="white");
              ax.set_yticks(pd.date_range("2024-1-1","2024-12-30",freq="YS"), range(2024,2024));
              ax.set_xlim(-7,7);
              for idx in range(len(setup)):
                  dt,product,level = setup["Date"][idx],setup["Milestone"][idx],setup["Level"][idx]
                  dt_str = dt.strftime("%b-%Y")
                  ax.annotate(dt_str+"\n"+product,xy=(0.1 if level>0 else -0.1,dt),
                              xytext=(level,dt),
                              arrowprops=dict(arrowstyle="-",color="red",linewidth=0.8),
                              va="center",fontsize=8);
                  ax.spines[["left", "top", "right", "bottom"]].set_visible(False);
                  ax.spines[["left"]].set_position(("axes", 0.5));
                  ax.xaxis.set_visible(False);                      
                  ax.set_title(f"{sector.title()} Milestones", pad=10, loc="left",
                                    fontsize=25, fontweight="bold");
                  ax.grid(False)
          plt.show() 

"""
-------------------
Profession Quadrant
-------------------
"""
class profession():
  def __init__(self,choice,file):
    self.choice = choice
    self.file = file
      
    print("\nMILESTONE\n")
    professionSides = {"MILESTONE":self.Milestone}
    profession_choice = input(f"{choice}-Side: ").upper()

    if profession_choice in professionSides:
        professionSides[profession_choice]()
    else:
        print("\nI recognize a dumbass when I see one...\n")

  def Milestone(self):
    # I. Transport data into proper format
    df = pd.read_excel(self.file,sheet_name=0,header=None)
    sector = self.choice
    datesOrigin = df.iloc[:,0].tolist()
    milestonesOrigin = df.iloc[:,1].tolist()
    print("\nNEW | VIEW\n")
    action = input("-> ")
    match str(action).upper():
        case "NEW":
          print("\nInput Date as: YYYY-MM-DD\n")
          date = str(input("-> "))
          try:
              datetime.strptime(date,"%Y-%m-%d")
          except:
              print("\nAre you retarded?\n")
              return
          print("\nWhat occurred that date?\n")
          milestoneInstance = input("-> ")
          milestoneAnnotation = [date,milestoneInstance]
          workbook = load_workbook(self.file)
          sheet_name = "MILESTONE"
          if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
          lastrowFirst = sheet.max_row
          lastrowSecond = sheet.max_row
          nextRow = max(lastrowFirst,lastrowSecond) + 1
          sheet.cell(row=nextRow,column=1,value=milestoneAnnotation[0])
          sheet.cell(row=nextRow,column=2,value=milestoneAnnotation[1])
          workbook.save(self.file)
            

        case "VIEW":
          df = pd.read_excel(self.file,sheet_name=0,header=None)
          dates = datesOrigin
          milestone = milestonesOrigin
          setup = pd.DataFrame(data={"Date": dates,"Milestone": milestone})
          setup["Date"] = pd.to_datetime(setup["Date"])
          setup["Level"] = [np.random.randint(-6,-2) if (i%2)==0 else np.random.randint(2,6) for i in
                                range(len(setup))]
          with plt.style.context("fivethirtyeight"):
              fig, ax = plt.subplots(figsize=(7,10))
              ax.plot([0,]* len(setup), setup.Date,"-o",color="black",markerfacecolor="white");
              ax.set_yticks(pd.date_range("2024-1-1","2024-12-30",freq="YS"), range(2024,2024));
              ax.set_xlim(-7,7);
              for idx in range(len(setup)):
                  dt,product,level = setup["Date"][idx],setup["Milestone"][idx],setup["Level"][idx]
                  dt_str = dt.strftime("%b-%Y")
                  ax.annotate(dt_str+"\n"+product,xy=(0.1 if level>0 else -0.1,dt),
                              xytext=(level,dt),
                              arrowprops=dict(arrowstyle="-",color="red",linewidth=0.8),
                              va="center",fontsize=8);
                  ax.spines[["left", "top", "right", "bottom"]].set_visible(False);
                  ax.spines[["left"]].set_position(("axes", 0.5));
                  ax.xaxis.set_visible(False);                      
                  ax.set_title(f"{sector.title()} Milestones", pad=10, loc="left",
                                    fontsize=25, fontweight="bold");
                  ax.grid(False)
          plt.show() 



"""
----------------
Finance Quadrant
----------------
"""
class finance():
  def __init__(self,choice,file):
    self.choice = choice
    self.file = file
      
    print("\nMILESTONE\n")
    financeSides = {"MILESTONE":self.Milestone}
    finance_choice = input(f"{choice}-Side: ").upper()

    if finance_choice in financeSides:
        financeSides[finance_choice]()
    else:
        print("\nI recognize a dumbass when I see one...\n")

  def Milestone(self):
    # I. Transport data into proper format
    df = pd.read_excel(self.file,sheet_name=0,header=None)
    sector = self.choice
    datesOrigin = df.iloc[:,0].tolist()
    milestonesOrigin = df.iloc[:,1].tolist()
    print("\nNEW | VIEW\n")
    action = input("-> ")
    match str(action).upper():
        case "NEW":
          print("\nInput Date as: YYYY-MM-DD\n")
          date = str(input("-> "))
          try:
              datetime.strptime(date,"%Y-%m-%d")
          except:
              print("\nAre you retarded?\n")
              return
          print("\nWhat occurred that date?\n")
          milestoneInstance = input("-> ")
          milestoneAnnotation = [date,milestoneInstance]
          workbook = load_workbook(self.file)
          sheet_name = "MILESTONE"
          if sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
          lastrowFirst = sheet.max_row
          lastrowSecond = sheet.max_row
          nextRow = max(lastrowFirst,lastrowSecond) + 1
          sheet.cell(row=nextRow,column=1,value=milestoneAnnotation[0])
          sheet.cell(row=nextRow,column=2,value=milestoneAnnotation[1])
          workbook.save(self.file)
            

        case "VIEW":
          df = pd.read_excel(self.file,sheet_name=0,header=None)
          dates = datesOrigin
          milestone = milestonesOrigin
          setup = pd.DataFrame(data={"Date": dates,"Milestone": milestone})
          setup["Date"] = pd.to_datetime(setup["Date"])
          setup["Level"] = [np.random.randint(-6,-2) if (i%2)==0 else np.random.randint(2,6) for i in
                                range(len(setup))]
          with plt.style.context("fivethirtyeight"):
              fig, ax = plt.subplots(figsize=(7,10))
              ax.plot([0,]* len(setup), setup.Date,"-o",color="black",markerfacecolor="white");
              ax.set_yticks(pd.date_range("2024-1-1","2024-12-30",freq="YS"), range(2024,2024));
              ax.set_xlim(-7,7);
              for idx in range(len(setup)):
                  dt,product,level = setup["Date"][idx],setup["Milestone"][idx],setup["Level"][idx]
                  dt_str = dt.strftime("%b-%Y")
                  ax.annotate(dt_str+"\n"+product,xy=(0.1 if level>0 else -0.1,dt),
                              xytext=(level,dt),
                              arrowprops=dict(arrowstyle="-",color="red",linewidth=0.8),
                              va="center",fontsize=8);
                  ax.spines[["left", "top", "right", "bottom"]].set_visible(False);
                  ax.spines[["left"]].set_position(("axes", 0.5));
                  ax.xaxis.set_visible(False);                      
                  ax.set_title(f"{sector.title()} Milestones", pad=10, loc="left",
                                    fontsize=25, fontweight="bold");
                  ax.grid(False)
          plt.show() 




if __name__ == "__main__":
    # File directory in your machine
    directory = "FOLDER CONTAINIGN ALL EXCELS"

    # Use if necessary:
    delete_ds_store_files(directory)
    
    # Allocates Excel Files into the program
    files = os.listdir(directory)
    sorted_files = sorted(files, key=lambda x: int(x[0]))
    file_dictionary = {}
    for i, file in enumerate(sorted_files):
        file_dictionary[f"file_{i+1}"] = os.path.join(directory, file)

    all_files = [i for i in file_dictionary.values()]
        
    # Outset    
    beginning(*all_files)























































