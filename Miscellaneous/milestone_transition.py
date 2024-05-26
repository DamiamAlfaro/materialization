import openpyxl
import json

source = "FILE"

with open(source,"r") as file: 
    read = file.read()

reshapingFile = json.loads(read)

datesMilestones = []

for i,j in zip(reshapingFile[0],reshapingFile[1]):
    litsNew = []
    litsNew.append(i)
    litsNew.append(j)
    datesMilestones.append(litsNew)

excel_file = "FILE"
workbook = openpyxl.load_workbook(excel_file)
sheet_name = "MILESTONE"
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

for row_index, row_data in enumerate(datesMilestones, start=1):
    for col_index, value in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)

workbook.save(excel_file)








