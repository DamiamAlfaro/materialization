import openpyxl
import pandas as pd


accuracy_file = "/Users/damiamalfaro/Desktop/Materialization/Transition/tth_math_accuracy.txt"
with open(accuracy_file,"r") as file:
    lines = file.readlines()

filter = [instance[:-1].split("=") for instance in lines]


dates = []
for instance in filter:
    new = instance[0].split(",")
    year = int(new[0][2:-1])
    month = int(new[1][2:-1])
    day = int(new[2][2:-3])
    date = [year,month,day,float(instance[1][1:])]
    dates.append(date)


df = pd.DataFrame(dates)

excel_file = "FILE"
workbook = openpyxl.load_workbook(excel_file)
sheet_name = "TESTING"
sheet_name = 'TESTING'
if sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]

for row_index, row_data in enumerate(dates, start=1):
    for col_index, value in enumerate(row_data, start=1):
        sheet.cell(row=row_index, column=col_index, value=value)

workbook.save(excel_file)










   
