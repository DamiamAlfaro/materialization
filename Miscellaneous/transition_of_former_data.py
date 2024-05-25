
import pandas as pd
from openpyxl import Workbook, load_workbook
import json
import re

with open("FILE SOURCE") as f1:
    lines = f1.readlines()



outsets = [line[1:-1].replace("'",'"') for line in lines if line[0] == "0"]
halts = [kine[1:-1].replace("'",'"') for kine in lines if kine[0] == "1"]

cleaned_outsets = [re.sub(r'(\d+):', r'"\1":', cle) for cle in outsets]
kleaned_halts = [re.sub(r'(\d+):', r'"\1":', kle) for kle in halts] 

beginnings = [json.loads(instance) for instance in cleaned_outsets]
endings = [json.loads(finals) for finals in kleaned_halts]

def removing_last(self):
    new_list = []
    for initial in self:
        last = list(initial.keys())[-1]
        initial.pop(last)
        new_list.append(initial)
    
    return new_list

initiations = [removing_last(beginnings)]
finals = [removing_last(endings)]

outset_keys = [[d[key] for key in d] for d in initiations[0]]
halt_keys = [[d[key] for key in d] for d in finals[0]]

print(outset_keys[0])

allocation = "FILE DESTINATION"

wb = load_workbook(allocation)
ws = wb["ACTION"]

# Write data to the sheet
for row_index, row_data in enumerate(halt_keys, start=1):
    for col_index, value in enumerate(row_data, start=1):
        cell = ws.cell(row=row_index, column=col_index)
        cell.value = value

# Save the workbook
wb.save(allocation)






































   
